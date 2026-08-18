"""
Shopify Order Reconciliation Builder — multi-store SaaS edition.

Usage:
    python reconcile.py --store KUDBI --inputs-dir ./april_inputs
    python reconcile.py --store KANTHA --inputs-dir ./inputs --month MAY --year 2026

Each store has a JSON config in ./stores/<STORE>.json containing:
  - product_costs:  parent_product_name -> per-piece INR cost

To onboard a new store, copy stores/_TEMPLATE.json to stores/<STORE>.json and edit.

Inputs (in --inputs-dir):
  shopify_orders.csv
  bluedart*.xlsx   (one or more)
  expressfly.csv
  delivery*.xlsx   (optional, one or more — Delhivery-style tracking export)

Output:
  /mnt/user-data/outputs/<MONTH>_<YEAR>_<STORE>_ORDERS.xlsx
"""

import argparse
import calendar
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ORDERS_COLUMNS = [
    "Name", "Financial Status", "Paid at", "Fulfillment Status", "Fulfilled at",
    "Subtotal", "Shipping", "Taxes", "Total",
    "Bluedart", "Other",
    "Lineitem name", "Created at", "Lineitem quantity", "Lineitem price",
    "Lineitem compare at price", "Lineitem sku", "Lineitem requires shipping",
    "Lineitem taxable", "Lineitem fulfillment status",
    "Billing Name", "Billing City", "Billing Zip", "Billing Province",
    "Billing Country", "Billing Phone",
    "Shipping Name", "Shipping Address1", "Shipping Address2", "Shipping City",
    "Shipping Zip", "Shipping Province", "Shipping Country", "Shipping Phone",
    "Notes", "Cancelled at", "Payment Method", "Payment Reference",
    "Refunded Amount", "Vendor", "Outstanding Balance",
    "Employee", "Location", "Device ID", "Id", "Tags", "Risk Level", "Source",
    "Lineitem discount",
    "Tax 1 Name", "Tax 1 Value", "Tax 2 Name", "Tax 2 Value",
    "Tax 3 Name", "Tax 3 Value", "Tax 4 Name", "Tax 4 Value",
    "Tax 5 Name", "Tax 5 Value",
    "Phone", "Receipt Number", "Duties",
    "Billing Province Name", "Shipping Province Name",
    "Payment ID", "Payment Terms Name", "Next Payment Due At", "Payment References",
]

YELLOW = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
HEADER_FONT = Font(name="Arial", size=11, bold=True)


# ----- config loader ---------------------------------------------- #

def load_store_config(store_name, stores_dir):
    cfg_path = Path(stores_dir) / f"{store_name}.json"
    if not cfg_path.exists():
        sys.exit(
            f"No config found at {cfg_path}\n"
            f"Create one by copying {Path(stores_dir) / '_TEMPLATE.json'}"
        )
    cfg = json.loads(cfg_path.read_text())
    # Sanitize: ignore underscore-keyed instruction entries
    cfg["product_costs"] = {
        k: v for k, v in cfg.get("product_costs", {}).items()
        if not k.startswith("_")
    }
    return cfg


def register_new_products(store_name, stores_dir, product_names):
    """
    Add any parent products not yet in stores/<STORE>.json with a null cost, so
    they show up as yellow (needs-a-cost) cells in the workbook. Existing
    entries — including underscore-keyed instructions — are left untouched.

    Returns the list of newly added product names.
    """
    cfg_path = Path(stores_dir) / f"{store_name}.json"
    raw = json.loads(cfg_path.read_text())
    costs = raw.setdefault("product_costs", {})
    new = [p for p in sorted(product_names) if p and p not in costs]
    if not new:
        return []
    for p in new:
        costs[p] = None
    cfg_path.write_text(json.dumps(raw, indent=2) + "\n")
    return new


# ----- helpers ---------------------------------------------------- #

def strip_hash(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    if s.startswith("#"):
        s = s[1:]
    if s.isdigit():
        return int(s)
    return s


ORDER_SUFFIX_RE = re.compile(r"^(\d+)-\S+$")


def normalize_order_no(v):
    """Collapse a courier's re-ship suffix back onto the base order number.

    Bluedart writes a re-shipped consignment as '79653-CLONE' (and
    '-CLONE-CLONE' on a second re-ship); the Other couriers use '-V1', '-03',
    '-A'. Shopify never suffixes its order names, so the suffixed form matches
    no order at all and the shipment goes uncounted.
    """
    n = strip_hash(v)
    if isinstance(n, str):
        m = ORDER_SUFFIX_RE.match(n.strip())
        if m:
            return int(m.group(1))
    return n


def _pickup_key(v):
    """Sortable pickup date; parse_date falls back to a raw string or None."""
    return v if isinstance(v, datetime) else datetime.min


def parent_product(lineitem_name):
    if lineitem_name is None or (isinstance(lineitem_name, float) and pd.isna(lineitem_name)):
        return None
    s = str(lineitem_name)
    return s.split(" - ")[0].strip() if " - " in s else s.strip()


def clean_str(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    return str(v).strip()


def parse_date(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # Excel serial date (delivery*.xlsx stores pickup date as a raw number)
        try:
            return from_excel(v)
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s


def detect_month_year(shopify_df):
    s = pd.to_datetime(shopify_df["Created at"], errors="coerce", utc=True)
    if s.notna().sum() == 0:
        s = pd.to_datetime(shopify_df["Fulfilled at"], errors="coerce", utc=True)
    s = s.dropna()
    if len(s) == 0:
        now = datetime.now()
        return now.month, now.year
    top = s.dt.tz_convert(None).dt.to_period("M").value_counts().index[0]
    return top.month, top.year


# ----- input readers ---------------------------------------------- #

def build_bluedart_status(bluedart_paths):
    rows = []
    seen = set()
    for p in bluedart_paths:
        wb = load_workbook(p, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = {clean_str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
        need = ["Reference No", "Status Description", "WayBill No", "P/U_Date", "Rto Reason"]
        idx = {k: header.get(k) for k in need}
        missing = [k for k, v in idx.items() if v is None]
        if missing:
            raise ValueError(f"{p} missing columns: {missing}")
        for r in range(2, ws.max_row + 1):
            ord_no = normalize_order_no(ws.cell(r, idx["Reference No"]).value)
            if ord_no is None:
                continue
            wb_no = ws.cell(r, idx["WayBill No"]).value
            key = (ord_no, wb_no)
            if key in seen:
                continue
            seen.add(key)
            rows.append([
                ord_no,
                clean_str(ws.cell(r, idx["Status Description"]).value),
                strip_hash(wb_no),
                parse_date(ws.cell(r, idx["P/U_Date"]).value),
                clean_str(ws.cell(r, idx["Rto Reason"]).value),
            ])
    return rows


def dedupe_other_rows(rows):
    """One row per order number for an Other-courier source.

    Re-ship suffixes collapse several rows onto the same order, so keeping
    whichever row was read first would let file order decide the status. Same
    rule as the Bluedart side: a definite outcome (Delivered / RTO) beats one
    still in flight, then the latest pickup wins.
    """
    best = {}
    for row in rows:
        n, status, date_v = row[0], row[1], row[4]
        definite = is_other_delivered(status) or is_other_rto(status)
        rank = (definite, _pickup_key(date_v))
        if n not in best or rank > best[n][0]:
            best[n] = (rank, row)
    return [row for _rank, row in best.values()]


def build_other_status(expressfly_csv):
    df = pd.read_csv(expressfly_csv, dtype=str)
    keep = ["Order Number", "Order Status", "Airwaybill Number", "Courier Name",
            "Order Date", "Service Provider Comment", "No of Attempt", "Zone"]
    df = df[keep].copy()
    df["Order Number"] = df["Order Number"].apply(normalize_order_no)
    df["Airwaybill Number"] = df["Airwaybill Number"].apply(strip_hash)
    df["No of Attempt"] = pd.to_numeric(df["No of Attempt"], errors="coerce").astype("Int64")
    df["Order Date"] = df["Order Date"].apply(parse_date)
    df = df.dropna(subset=["Order Number"])
    return dedupe_other_rows(df.values.tolist())


def build_delivery_status(delivery_paths):
    """
    Read delivery*.xlsx (Delhivery-style export) into the same row shape as
    build_other_status():
      [Order Number, Order Status, Airwaybill Number, Courier Name,
       Order Date, Service Provider Comment, No of Attempt, Zone]

    Columns used: 'order #', 'current status' ('Delivered' / 'RTO' / ...),
    'waybill number', 'pickup date', 'last scan remarks', 'attempt count'.
    """
    rows = []
    for p in delivery_paths:
        wb = load_workbook(p, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = {}
        for c in range(1, ws.max_column + 1):
            h = clean_str(ws.cell(1, c).value)
            if h:
                header[h.lower()] = c
        need = ["order #", "current status", "waybill number"]
        missing = [k for k in need if k not in header]
        if missing:
            raise ValueError(f"{p} missing columns: {missing}")

        def cell(r, key):
            c = header.get(key)
            return ws.cell(r, c).value if c else None

        for r in range(2, ws.max_row + 1):
            ord_no = normalize_order_no(cell(r, "order #"))
            if ord_no is None:
                continue
            attempts = pd.to_numeric(cell(r, "attempt count"), errors="coerce")
            rows.append([
                ord_no,
                clean_str(cell(r, "current status")),
                strip_hash(cell(r, "waybill number")),
                "DELHIVERY",
                parse_date(cell(r, "pickup date")),
                clean_str(cell(r, "last scan remarks")),
                None if pd.isna(attempts) else int(attempts),
                None,
            ])
    return dedupe_other_rows(rows)


def merge_other_status(expressfly_rows, delivery_rows):
    """
    Combine expressfly + delivery rows, one row per order number.
    A definite status (Delivered / RTO) beats an in-flight one, otherwise the
    first source seen (expressfly) wins.
    """
    def definite(status):
        return is_other_delivered(status) or is_other_rto(status)

    merged = {}
    order = []
    for row in list(expressfly_rows) + list(delivery_rows):
        n = row[0]
        if n not in merged:
            merged[n] = row
            order.append(n)
        elif definite(row[1]) and not definite(merged[n][1]):
            merged[n] = row
    return [merged[n] for n in order]


def build_orders(shopify_csv):
    """
    Returns (orders_df_deduped, raw_fulfilled_df, totals).
      orders_df_deduped:  one row per fulfilled order, shaped to ORDERS_COLUMNS
      raw_fulfilled_df:   all line items for fulfilled orders (for ProductCost calc)
      totals:             dict with total_sales_raw and after_cancellation_sales
    """
    df = pd.read_csv(shopify_csv, dtype=str)
    df["Name"] = df["Name"].apply(strip_hash)
    df = df.dropna(subset=["Name"])

    # Total Sales (raw) = sum of Total across all orders BEFORE removing unfulfilled.
    # Total appears only on the first line of each order.
    raw_order_first = df.drop_duplicates(subset=["Name"], keep="first")
    total_sales_raw = pd.to_numeric(raw_order_first["Total"], errors="coerce").fillna(0).sum()

    # Determine fulfilment per order.
    order_status = df.groupby("Name")["Fulfillment Status"].apply(
        lambda s: next((v for v in s if isinstance(v, str) and v.strip()), None)
    )
    fulfilled = set(order_status[order_status == "fulfilled"].index)
    df_full = df[df["Name"].isin(fulfilled)].copy()
    df_dedup = df_full.drop_duplicates(subset=["Name"], keep="first").copy()

    after_cancel_sales = pd.to_numeric(df_dedup["Total"], errors="coerce").fillna(0).sum()

    out = pd.DataFrame({c: df_dedup.get(c) for c in ORDERS_COLUMNS})
    out["Bluedart"] = None
    out["Other"] = None

    numeric_cols = ("Subtotal", "Shipping", "Taxes", "Total",
                    "Lineitem price", "Lineitem compare at price", "Lineitem discount",
                    "Lineitem quantity", "Refunded Amount", "Outstanding Balance",
                    "Tax 1 Value", "Tax 2 Value", "Tax 3 Value", "Tax 4 Value",
                    "Tax 5 Value", "Duties")
    for c in numeric_cols:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    totals = {
        "total_sales_raw": round(float(total_sales_raw), 2),
        "after_cancellation_sales": round(float(after_cancel_sales), 2),
    }
    return out, df_full, totals


# ----- derived sheets --------------------------------------------- #

# Status classification
def is_bluedart_delivered(status):
    return status == "SHIPMENT DELIVERED"

# Bluedart statuses that are terminal returns but contain neither "RTO" nor
# "RETURN" in the text, so the substring test below would miss them.
BLUEDART_RTO_STATUSES = {
    "DELIVERED BACK TO SHIPPER",
}

def is_bluedart_rto(status):
    if not status:
        return False
    s = str(status).upper()
    if s in BLUEDART_RTO_STATUSES:
        return True
    return "RTO" in s or "RETURN" in s

def is_other_delivered(status):
    return status == "Delivered"

def is_other_rto(status):
    return status and "RTO" in str(status)


def resolve_bluedart_status(bd_rows):
    """One Bluedart status per order number.

    A re-ship leaves several rows on the same order once '-CLONE' references
    collapse onto their base. Picking the last row read would make the answer
    depend on file order, so resolve it: a definite outcome (Delivered / RTO)
    beats one still in flight, and among rows of equal standing the latest
    pickup wins — the order takes the fate of its final shipment.
    """
    best = {}
    for row in bd_rows:
        n, status, _wb, pu = row[0], row[1], row[2], row[3]
        definite = is_bluedart_delivered(status) or is_bluedart_rto(status)
        rank = (definite, _pickup_key(pu))
        if n not in best or rank > best[n][0]:
            best[n] = (rank, status)
    return {n: status for n, (_rank, status) in best.items()}


def classify_orders(orders_df, bd_rows, other_rows, product_costs):
    """
    Returns a dict of classified order sets and product-level rollup data.

    NEW RULE: For an order in orders_df (i.e. fulfilled) where:
      - Financial Status == 'paid'
      - AND the order is NOT found in Bluedart OR Other status sheets
    => treat it as Other-Delivered (i.e. assumed delivered).
    """
    bd_index = resolve_bluedart_status(bd_rows)
    other_index = {r[0]: r[1] for r in other_rows}

    bd_delivered = set()
    bd_rto = set()
    other_delivered = set()
    other_rto = set()
    unknown = set()
    paid_no_status_assumed_delivered = set()

    fulfilled_names = set(orders_df["Name"].dropna().tolist())
    paid_orders = set(orders_df.loc[orders_df["Financial Status"] == "paid", "Name"].dropna().tolist())

    for n in fulfilled_names:
        bd_status = bd_index.get(n)
        ot_status = other_index.get(n)

        if is_bluedart_delivered(bd_status):
            bd_delivered.add(n)
        elif is_bluedart_rto(bd_status):
            bd_rto.add(n)
        elif is_other_delivered(ot_status):
            other_delivered.add(n)
        elif is_other_rto(ot_status):
            other_rto.add(n)
        else:
            # No definite status in either source.
            # If paid -> assume delivered (Other column).
            if n in paid_orders:
                paid_no_status_assumed_delivered.add(n)
                other_delivered.add(n)
            else:
                # COD order still in flight (or with no record at all). It is
                # neither delivered nor RTO yet, so it gets its own bucket
                # instead of silently vanishing from the percentages.
                unknown.add(n)

    return {
        "bd_delivered": bd_delivered,
        "bd_rto": bd_rto,
        "other_delivered": other_delivered,
        "other_rto": other_rto,
        "unknown": unknown,
        "paid_no_status_assumed_delivered": paid_no_status_assumed_delivered,
    }


def build_payment_split(orders_df, classification):
    """Delivered / RTO / Unknown order counts split by payment mode.

    Prepaid = Shopify 'Financial Status' == 'paid'; everything else is COD
    (Shopify leaves COD orders as 'pending' until the courier remits). RTO
    behaves so differently across the two that a blended RTO% hides the only
    number worth acting on.
    """
    paid = set(orders_df.loc[orders_df["Financial Status"] == "paid",
                             "Name"].dropna().tolist())
    delivered = classification["bd_delivered"] | classification["other_delivered"]
    rto = classification["bd_rto"] | classification["other_rto"]
    unknown = classification["unknown"]

    split = {}
    for label, in_mode in (("Prepaid", lambda n: n in paid),
                           ("COD", lambda n: n not in paid)):
        split[label] = {
            "delivered": sum(1 for n in delivered if in_mode(n)),
            "rto": sum(1 for n in rto if in_mode(n)),
            "unknown": sum(1 for n in unknown if in_mode(n)),
        }
    return split


def build_product_cost(raw_fulfilled, classification, product_costs):
    """ProductCost data: variant counts (Bluedart + Other delivered) and parent rollup."""
    bd_delivered = classification["bd_delivered"]
    other_delivered = classification["other_delivered"]

    bd_variant = defaultdict(int)
    other_variant = defaultdict(int)
    parent_bd = defaultdict(int)
    parent_other = defaultdict(int)

    for _, row in raw_fulfilled.iterrows():
        name = row.get("Name")
        variant = row.get("Lineitem name")
        try:
            qty = int(row.get("Lineitem quantity") or 0)
        except (ValueError, TypeError):
            qty = 0
        if not variant or qty == 0:
            continue
        parent = parent_product(variant)
        if name in bd_delivered:
            bd_variant[variant] += qty
            parent_bd[parent] += qty
        elif name in other_delivered:
            other_variant[variant] += qty
            parent_other[parent] += qty

    parents = sorted(set(parent_bd) | set(parent_other))
    rollup = [
        (p, parent_bd.get(p, 0), parent_other.get(p, 0), product_costs.get(p))
        for p in parents
    ]
    return {
        "bd_variants": sorted(bd_variant.items()),
        "other_variants": sorted(other_variant.items()),
        "rollup": rollup,
    }


def build_product_pnl(raw_fulfilled, classification, product_costs):
    """
    Product-wise profit/loss based on DELIVERED orders only (Bluedart + Other
    delivered). RTO units don't generate revenue and their stock returns, so
    they're excluded from both sales and product cost.

    Per parent product returns:
      (name, total_qty, delivered_qty, sales, cost_per_piece,
       paid_orders, delivered_orders)
      - total_qty       : units across all fulfilled orders (delivered + RTO etc.)
      - delivered_qty   : units delivered
      - sales           : sum(Lineitem price * qty - Lineitem discount) for
                          delivered line items
      - cost_per_piece  : from store config (None if not configured)
      - paid_orders     : distinct delivered orders for this product that are
                          prepaid (Financial Status == 'paid')
      - delivered_orders: distinct delivered orders containing this product
      - total_orders    : distinct fulfilled orders containing this product
    """
    delivered = classification["bd_delivered"] | classification["other_delivered"]

    def to_num(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    total_qty_by_parent = defaultdict(int)
    qty_by_parent = defaultdict(int)
    sales_by_parent = defaultdict(float)
    orders_by_parent = defaultdict(set)
    total_orders_by_parent = defaultdict(set)
    paid_orders_by_parent = defaultdict(set)

    for _, row in raw_fulfilled.iterrows():
        variant = row.get("Lineitem name")
        if not variant:
            continue
        qty = int(to_num(row.get("Lineitem quantity")))
        if qty == 0:
            continue
        parent = parent_product(variant)
        total_qty_by_parent[parent] += qty
        name = row.get("Name")
        if name is not None:
            total_orders_by_parent[parent].add(name)

        if name not in delivered:
            continue
        price = to_num(row.get("Lineitem price"))
        disc = to_num(row.get("Lineitem discount"))
        qty_by_parent[parent] += qty
        sales_by_parent[parent] += price * qty - disc
        orders_by_parent[parent].add(name)
        if row.get("Financial Status") == "paid":
            paid_orders_by_parent[parent].add(name)

    parents = sorted(set(total_qty_by_parent) | set(qty_by_parent) | set(sales_by_parent))
    return [
        (p, total_qty_by_parent.get(p, 0), qty_by_parent.get(p, 0),
         round(sales_by_parent.get(p, 0.0), 2), product_costs.get(p),
         len(paid_orders_by_parent.get(p, set())),
         len(orders_by_parent.get(p, set())),
         len(total_orders_by_parent.get(p, set())))
        for p in parents
    ]


# ----- writers ---------------------------------------------------- #

def write_orders(ws, orders_df, classification):
    ws.append(ORDERS_COLUMNS)
    for cell in ws[1]:
        cell.font = HEADER_FONT

    assumed = classification["paid_no_status_assumed_delivered"]

    for i, row in enumerate(orders_df.itertuples(index=False), start=2):
        name_val = row[0]
        for j, val in enumerate(row, start=1):
            if pd.isna(val):
                val = None
            ws.cell(i, j, val)

        # Bluedart col (J=10): VLOOKUP
        ws.cell(i, 10, f'=IFERROR(VLOOKUP(A{i},BluedartSatus!$A:$B,2,FALSE),"")')

        # Other col (K=11): VLOOKUP, but if the order is in the "paid + no status"
        # set, hardcode "Delivered" because there's nothing to look up.
        if name_val in assumed:
            ws.cell(i, 11, "Delivered")
        else:
            ws.cell(i, 11, f'=IFERROR(VLOOKUP(A{i},OtherStatus!$A:$B,2,FALSE),"")')

    widths = {1: 8, 2: 14, 3: 18, 4: 16, 5: 18, 6: 9, 7: 9, 8: 8, 9: 10,
              10: 22, 11: 22, 12: 36}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def write_bluedart_status(ws, rows):
    ws.append(["Order Number", "Status Description", "WayBill No", "P/U_Date", "Rto Reason"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 22
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 4).number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A2"


def write_other_status(ws, rows, assumed_orders):
    """
    Append rows from expressfly, then append synthetic rows for
    paid-orders-with-no-status (so that the VLOOKUP in Orders would still
    work if we ever switched back from hardcoded "Delivered").
    """
    ws.append(["Order Number", "Order Status", "Airwaybill Number", "Courier Name",
               "Order Date", "Service Provider Comment", "No of Attempt", "Zone"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    existing = set()
    for row in rows:
        ws.append(row)
        existing.add(row[0])
    # Synthetic rows for assumed-delivered
    for n in sorted(assumed_orders):
        if n in existing:
            continue
        ws.append([n, "Delivered", None, "ASSUMED (paid, no tracking)", None,
                   "Auto-marked delivered: paid order with no delivery record", None, None])
    widths = [14, 18, 22, 22, 14, 36, 12, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_product_cost(ws, data):
    bd = data["bd_variants"]
    other = data["other_variants"]
    rollup = data["rollup"]

    ws["A1"] = "Bluedart Variant"
    ws["B1"] = "Qty"
    ws["D1"] = "Other Variant"
    ws["E1"] = "Qty"
    ws["G2"] = "Product Name"
    ws["H2"] = "Bluedart"
    ws["I2"] = "Other"
    ws["J2"] = "Total"
    ws["K2"] = "Cost"
    ws["L2"] = "Total Cost"

    for ref in ("A1", "B1", "D1", "E1", "G2", "H2", "I2", "J2", "K2", "L2"):
        ws[ref].font = HEADER_FONT

    for i, (variant, qty) in enumerate(bd, start=2):
        ws.cell(i, 1, variant)
        ws.cell(i, 2, qty)
    for i, (variant, qty) in enumerate(other, start=2):
        ws.cell(i, 4, variant)
        ws.cell(i, 5, qty)

    start = 3
    for i, (pname, bd_qty, other_qty, cost) in enumerate(rollup):
        r = start + i
        ws.cell(r, 7, pname)
        ws.cell(r, 8, bd_qty)
        ws.cell(r, 9, other_qty)
        ws.cell(r, 10, f"=H{r}+I{r}")
        if cost is None:
            ws.cell(r, 11, None).fill = YELLOW
        else:
            ws.cell(r, 11, cost)
        ws.cell(r, 12, f"=J{r}*K{r}")

    if rollup:
        total_row = start + len(rollup)
        last = total_row - 1
        ws.cell(total_row, 7, "TOTAL").font = HEADER_FONT
        ws.cell(total_row, 8, f"=SUM(H{start}:H{last})").font = HEADER_FONT
        ws.cell(total_row, 9, f"=SUM(I{start}:I{last})").font = HEADER_FONT
        ws.cell(total_row, 10, f"=SUM(J{start}:J{last})").font = HEADER_FONT
        ws.cell(total_row, 12, f"=SUM(L{start}:L{last})").font = HEADER_FONT
        ws.cell(total_row, 13, f"=IFERROR(L{total_row}/J{total_row},0)").font = HEADER_FONT

    widths = {1: 55, 2: 8, 4: 55, 5: 8, 7: 55, 8: 10, 9: 10, 10: 10,
              11: 10, 12: 14, 13: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    return start + len(rollup) if rollup else start  # total row index


def write_pnl(ws, pnl_rows):
    """Product-wise profit/loss sheet (delivered orders only).

    Shipping is pulled live from the Reconciliation tab:
      - Shipping per product = Delivered Qty * Reconciliation!E23 (avg ship/delivered)
    Meta spend per product is entered manually (yellow column, excl GST); an
    18% GST is added automatically in the next column.
    Profit/Loss = Sales - Product Cost - Shipping - Meta(+18% GST).
    Also shows prepaid (paid) delivered orders per product and the paid %.
    """
    headers = ["Product Name", "Total Qty", "Delivered Qty", "Sales", "Cost/pc",
               "Product Cost", "Shipping", "Meta Spend", "Meta +18% GST",
               "Profit/Loss", "Margin %", "Paid Orders",
               "Paid % (Delivered)", "Paid % (Total)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h).font = HEADER_FONT

    start = 2
    total_row = start + len(pnl_rows)
    tot_paid = tot_delivered_orders = tot_orders = 0
    for i, (pname, total_qty, qty, sales, cost, paid_orders,
            delivered_orders, total_orders) in enumerate(pnl_rows):
        r = start + i
        tot_paid += paid_orders
        tot_delivered_orders += delivered_orders
        tot_orders += total_orders
        ws.cell(r, 1, pname)
        ws.cell(r, 2, total_qty)                       # Total Qty (all fulfilled)
        ws.cell(r, 3, qty)                             # Delivered Qty
        ws.cell(r, 4, sales)
        if cost is None:
            ws.cell(r, 5, None).fill = YELLOW
        else:
            ws.cell(r, 5, cost)
        ws.cell(r, 6, f"=C{r}*E{r}")                  # Product Cost
        ws.cell(r, 7, f"=C{r}*Reconciliation!$E$23")  # Shipping
        ws.cell(r, 8, None).fill = YELLOW             # Meta Spend (manual, excl GST)
        ws.cell(r, 9, f"=H{r}*1.18")                  # Meta +18% GST
        ws.cell(r, 10, f"=D{r}-F{r}-G{r}-I{r}")       # Profit/Loss
        ws.cell(r, 11, f"=IFERROR(J{r}/D{r}*100,0)")  # Margin %
        ws.cell(r, 12, paid_orders)                   # Paid Orders
        pct_del = round(paid_orders / delivered_orders * 100, 2) if delivered_orders else 0
        pct_tot = round(paid_orders / total_orders * 100, 2) if total_orders else 0
        ws.cell(r, 13, pct_del)                       # Paid % of delivered orders
        ws.cell(r, 14, pct_tot)                       # Paid % of total orders

    if pnl_rows:
        last = total_row - 1
        ws.cell(total_row, 1, "TOTAL").font = HEADER_FONT
        for col in (2, 3, 4, 6, 7, 8, 9, 10, 12):
            c = get_column_letter(col)
            ws.cell(total_row, col, f"=SUM({c}{start}:{c}{last})").font = HEADER_FONT
        ws.cell(total_row, 11,
                f"=IFERROR(J{total_row}/D{total_row}*100,0)").font = HEADER_FONT
        tot_pct_del = round(tot_paid / tot_delivered_orders * 100, 2) if tot_delivered_orders else 0
        tot_pct_tot = round(tot_paid / tot_orders * 100, 2) if tot_orders else 0
        ws.cell(total_row, 13, tot_pct_del).font = HEADER_FONT
        ws.cell(total_row, 14, tot_pct_tot).font = HEADER_FONT

    widths = {1: 55, 2: 12, 3: 14, 4: 16, 5: 10, 6: 16, 7: 16, 8: 16, 9: 16,
              10: 16, 11: 12, 12: 12, 13: 18, 14: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B2"


def write_delivery_analysis(ws, rollup_len, pc_start=3):
    """Unit-level delivery split per parent product.

    RTO% and Unknown% are measured against Total Qty from their own unit
    counts, not derived as 100-Delivery%. Delivery% + RTO% + Unknown% == 100
    only because the three unit buckets partition Total Qty; column L makes
    that visible instead of assuming it.
    """
    ws["B1"] = "Data derived based on delivered (quantities / units)"
    ws["B1"].font = HEADER_FONT
    headers = ["Product Name", "Total Qty", "Bluedart Qty", "Other Qty",
               "Delivered Qty", "RTO Qty", "Unknown Qty",
               "Delivery%", "RTO%", "Unknown%", "Check"]
    for i, h in enumerate(headers, start=2):
        ws.cell(2, i, h).font = HEADER_FONT

    for i in range(rollup_len):
        r = 3 + i
        pc_r = pc_start + i
        ws.cell(r, 2, f"=ProductCost!G{pc_r}")
        ws.cell(r, 3, None)  # C Total Qty   - filled by populate_unit_counts
        ws.cell(r, 4, f"=ProductCost!H{pc_r}")
        ws.cell(r, 5, f"=ProductCost!I{pc_r}")
        ws.cell(r, 6, f"=SUM(D{r}:E{r})")
        ws.cell(r, 7, None)  # G RTO Qty     - filled by populate_unit_counts
        ws.cell(r, 8, None)  # H Unknown Qty - filled by populate_unit_counts
        ws.cell(r, 9, f"=IFERROR(F{r}/C{r}*100,0)")
        ws.cell(r, 10, f"=IFERROR(G{r}/C{r}*100,0)")
        ws.cell(r, 11, f"=IFERROR(H{r}/C{r}*100,0)")
        ws.cell(r, 12, f"=I{r}+J{r}+K{r}")

    total_row = 3 + rollup_len
    if rollup_len:
        last = total_row - 1
        ws.cell(total_row, 2, "TOTAL").font = HEADER_FONT
        for col in (3, 4, 5, 7, 8):
            c = get_column_letter(col)
            ws.cell(total_row, col, f"=SUM({c}3:{c}{last})").font = HEADER_FONT
        ws.cell(total_row, 6, f"=SUM(D{total_row}:E{total_row})").font = HEADER_FONT
        ws.cell(total_row, 9,
                f"=IFERROR(F{total_row}/C{total_row}*100,0)").font = HEADER_FONT
        ws.cell(total_row, 10,
                f"=IFERROR(G{total_row}/C{total_row}*100,0)").font = HEADER_FONT
        ws.cell(total_row, 11,
                f"=IFERROR(H{total_row}/C{total_row}*100,0)").font = HEADER_FONT
        ws.cell(total_row, 12,
                f"=I{total_row}+J{total_row}+K{total_row}").font = HEADER_FONT

    # ----- Orders vs Units reconciliation note -----
    # DeliveryAnalysis counts UNITS; Reconciliation counts ORDERS. The gap is
    # delivered orders containing more than one unit.
    note_row = total_row + 2
    ws.cell(note_row, 2, "Delivered Units (this sheet)").font = HEADER_FONT
    ws.cell(note_row, 6, f"=F{total_row}" if rollup_len else 0).font = HEADER_FONT
    ws.cell(note_row + 1, 2, "Delivered Orders (Reconciliation)").font = HEADER_FONT
    ws.cell(note_row + 1, 6, "=Reconciliation!E6+Reconciliation!E7").font = HEADER_FONT
    ws.cell(note_row + 2, 2, "Difference (multi-unit orders)").font = HEADER_FONT
    ws.cell(note_row + 2, 6, f"=F{note_row}-F{note_row + 1}").font = HEADER_FONT

    widths = {2: 55, 3: 12, 4: 12, 5: 12, 6: 16, 7: 12, 8: 14,
              9: 12, 10: 10, 11: 12, 12: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C3"


def populate_unit_counts(ws_da, raw_fulfilled, rollup_products, classification):
    """Fill DeliveryAnalysis C (Total Qty), G (RTO Qty) and H (Unknown Qty).

    Counted from the same line items and the same order buckets that feed
    ProductCost columns H/I, so Delivered + RTO + Unknown reconciles to Total
    per parent product.
    """
    rto = classification["bd_rto"] | classification["other_rto"]
    unknown = classification["unknown"]

    parent_qty = defaultdict(int)
    parent_rto = defaultdict(int)
    parent_unknown = defaultdict(int)

    for _, row in raw_fulfilled.iterrows():
        name = row.get("Name")
        variant = row.get("Lineitem name")
        try:
            qty = int(row.get("Lineitem quantity") or 0)
        except (ValueError, TypeError):
            qty = 0
        if not variant or qty == 0:
            continue
        parent = parent_product(variant)
        parent_qty[parent] += qty
        if name in rto:
            parent_rto[parent] += qty
        elif name in unknown:
            parent_unknown[parent] += qty

    for i, pname in enumerate(rollup_products):
        r = 3 + i
        ws_da.cell(r, 3, parent_qty.get(pname, 0))
        ws_da.cell(r, 7, parent_rto.get(pname, 0))
        ws_da.cell(r, 8, parent_unknown.get(pname, 0))

    # A parent with no delivered units never enters the ProductCost rollup, so
    # its RTO/Unknown units would be invisible on this sheet. Surface it rather
    # than letting the columns silently disagree with Total Qty.
    orphans = {p: q for p, q in parent_qty.items() if p not in set(rollup_products)}
    if orphans:
        print("  WARNING: parents with no delivered units, omitted from "
              f"DeliveryAnalysis: {orphans}")


def write_payment_split(ws, payment_split, start_row=35):
    """Prepaid vs COD delivery/RTO block.

    Percentages are on the settled base (delivered + RTO) so the in-flight
    Unknown column cannot flatter either mode.
    """
    ws.cell(start_row, 4, "Payment Mode Split").font = HEADER_FONT

    hdr = start_row + 1
    headers = ["Orders", "Delivered", "RTO", "Unknown", "Settled",
               "Delivery% (settled)", "RTO% (settled)", "Share of Orders%"]
    for i, h in enumerate(headers, start=5):
        ws.cell(hdr, i, h).font = HEADER_FONT

    for i, label in enumerate(("Prepaid", "COD")):
        r = hdr + 1 + i
        d = payment_split[label]
        ws.cell(r, 4, label)
        ws.cell(r, 5, f"=SUM(F{r}:H{r})")   # Orders
        ws.cell(r, 6, d["delivered"])
        ws.cell(r, 7, d["rto"])
        ws.cell(r, 8, d["unknown"])
        ws.cell(r, 9, f"=F{r}+G{r}")        # Settled
        ws.cell(r, 10, f"=IFERROR(F{r}/I{r}*100,0)")
        ws.cell(r, 11, f"=IFERROR(G{r}/I{r}*100,0)")
        ws.cell(r, 12, "=IFERROR(E{r}/$E$4*100,0)".format(r=r))

    tot = hdr + 3
    ws.cell(tot, 4, "TOTAL").font = HEADER_FONT
    for col in range(5, 10):
        c = get_column_letter(col)
        ws.cell(tot, col, f"=SUM({c}{hdr + 1}:{c}{tot - 1})").font = HEADER_FONT
    ws.cell(tot, 10, f"=IFERROR(F{tot}/I{tot}*100,0)").font = HEADER_FONT
    ws.cell(tot, 11, f"=IFERROR(G{tot}/I{tot}*100,0)").font = HEADER_FONT
    ws.cell(tot, 12, f"=IFERROR(E{tot}/$E$4*100,0)").font = HEADER_FONT

    # TOTAL here must equal Fullfield Orders; if it does not, an order escaped
    # every bucket and the percentages above are being computed on a short base.
    ws.cell(tot + 1, 4, "Check vs Fullfield Orders").font = HEADER_FONT
    ws.cell(tot + 1, 5, f"=E{tot}-E4").font = HEADER_FONT

    for col, w in {5: 12, 6: 12, 7: 10, 8: 12, 9: 12,
                   10: 20, 11: 18, 12: 18}.items():
        cur = ws.column_dimensions[get_column_letter(col)].width or 0
        ws.column_dimensions[get_column_letter(col)].width = max(cur, w)


def write_reconciliation(ws, totals, totals_cls, orders_df, pc_total_row,
                         payment_split):
    """
    Reconciliation dashboard with the new auto-filled fields:
      I24 = total_sales_raw  (was manual)
      I25 = after_cancellation_sales  (was manual)
    """
    total_orders = totals_cls["total_orders"]
    fulfilled_orders = totals_cls["fulfilled_orders"]
    bd_delivered = totals_cls["bd_delivered"]
    bd_rto = totals_cls["bd_rto"]
    other_delivered = totals_cls["other_delivered"]
    other_rto = totals_cls["other_rto"]
    unknown = totals_cls["unknown"]
    unknown_amt = totals_cls["unknown_amt"]
    bd_delivered_amt = totals_cls["bd_delivered_amt"]
    bd_rto_amt = totals_cls["bd_rto_amt"]
    other_delivered_amt = totals_cls["other_delivered_amt"]
    other_rto_amt = totals_cls["other_rto_amt"]
    prepaid_count = totals_cls["prepaid_count"]
    prepaid_amt = totals_cls["prepaid_amt"]

    # ----- Left block -----
    ws["D3"] = "Total Orders"; ws["E3"] = total_orders
    ws["D4"] = "Fullfield Orders"; ws["E4"] = fulfilled_orders

    ws["D6"] = "Delivered"
    ws["E6"] = bd_delivered
    ws["F6"] = bd_delivered_amt
    ws["D7"] = "Other Delivered"
    ws["E7"] = other_delivered
    ws["F7"] = other_delivered_amt
    ws["D8"] = "RTO"
    ws["E8"] = bd_rto
    ws["F8"] = bd_rto_amt
    ws["D9"] = "Other RTO"
    ws["E9"] = other_rto
    ws["F9"] = other_rto_amt
    # Fulfilled orders that are neither delivered nor RTO: still in transit, not
    # picked up, or with no courier record at all. Without this row E6:E9 falls
    # short of E4 and Delivery% + RTO% does not add up to 100.
    ws["D10"] = "Unknown / In Transit"
    ws["E10"] = unknown
    ws["F10"] = unknown_amt

    ws["E11"] = "=SUM(E6:E10)"
    ws["F11"] = "=SUM(F6:F10)"

    ws["D13"] = "Prepaid orders"
    ws["E13"] = prepaid_count
    ws["F13"] = prepaid_amt

    ws["D14"] = "Other Courier%"; ws["E14"] = "=IFERROR((E7+E9)/E4*100,0)"
    ws["D15"] = "Bluedart%";       ws["E15"] = "=IFERROR((E8+E6)/E4*100,0)"
    ws["D16"] = "Total Delivery%"; ws["E16"] = "=IFERROR((E6+E7)/E4*100,0)"
    ws["D17"] = "Total RTO%";      ws["E17"] = "=IFERROR((E8+E9)/E4*100,0)"
    ws["D18"] = "Bluedart RTO%";   ws["E18"] = "=IFERROR(E8/(E8+E6)*100,0)"
    ws["D19"] = "Other RTO%";      ws["E19"] = "=IFERROR(E9/(E7+E9)*100,0)"
    ws["D20"] = "Prepaid orders";  ws["E20"] = "=IFERROR(E13/E3*100,0)"
    ws["D21"] = "Fullfillment%";   ws["E21"] = "=IFERROR(E4/E3*100,0)"
    ws["D22"] = "Cancelletion%";   ws["E22"] = "=100-E21"
    ws["D23"] = "Shiping Avg on Delivered";        ws["E23"] = "=IFERROR(-(I8+I9)/(E6+E7),0)"
    ws["D24"] = "Bluedart Avg on Delivered";       ws["E24"] = "=IFERROR(-I8/E6,0)"
    ws["D25"] = "Other Avg on Delivered";          ws["E25"] = "=IFERROR(-I9/E7,0)"
    ws["D26"] = "Total Shiping Avg";               ws["E26"] = "=IFERROR(-(I8+I9)/(E6+E7+E8+E9),0)"
    ws["D27"] = "Bluedart Total Shiping Avg";      ws["E27"] = "=IFERROR(-I8/(E6+E8),0)"
    ws["D28"] = "Other Total Shiping Avg";         ws["E28"] = "=IFERROR(-I9/(E7+E9),0)"
    # Completes the Delivery% / RTO% split: E16 + E17 + E29 = 100.
    ws["D29"] = "Unknown / In Transit%"; ws["E29"] = "=IFERROR(E10/E4*100,0)"
    ws["D30"] = "Status Coverage Check"; ws["E30"] = "=E16+E17+E29"

    # ----- Settled basis -----
    # E16/E17 divide by every fulfilled order, including those still in flight,
    # which drags RTO% down by however much was undelivered at export time.
    # Dividing by settled orders only makes months comparable.
    ws["D31"] = "Settled Orders (excl Unknown)"; ws["E31"] = "=E6+E7+E8+E9"
    ws["D32"] = "Delivery% (settled)"; ws["E32"] = "=IFERROR((E6+E7)/E31*100,0)"
    ws["D33"] = "RTO% (settled)";      ws["E33"] = "=IFERROR((E8+E9)/E31*100,0)"

    # ----- Payment mode split -----
    write_payment_split(ws, payment_split, start_row=35)

    # ----- Middle block (P&L) -----
    ws["H6"]  = "Net Sales";          ws["I6"]  = "=F6+F7"
    ws["H7"]  = "Meta Spend";         ws["I7"]  = None; ws["I7"].fill  = YELLOW
    ws["H8"]  = "Bluedart Shipping";  ws["I8"]  = None; ws["I8"].fill  = YELLOW
    ws["H9"]  = "Other Shipping";     ws["I9"]  = None; ws["I9"].fill  = YELLOW
    ws["H10"] = "Product Cost";       ws["I10"] = f"=-ProductCost!L{pc_total_row}"
    ws["H11"] = "Refunds";            ws["I11"] = None; ws["I11"].fill = YELLOW
    ws["H12"] = "Exchange";           ws["I12"] = None; ws["I12"].fill = YELLOW
    ws["H13"] = "Meta agency fees";   ws["I13"] = "=-(-(I7)*7)/100"
    ws["H14"] = "Rent";               ws["I14"] = None; ws["I14"].fill = YELLOW
    ws["H15"] = "Lightbill";          ws["I15"] = None; ws["I15"].fill = YELLOW
    ws["H16"] = "Salary";             ws["I16"] = None; ws["I16"].fill = YELLOW
    ws["K16"] = "Avg per order"
    ws["H17"] = "Gross";              ws["I17"] = "=SUM(I6:I16)"
    ws["J17"] = "=E6+E7";             ws["K17"] = "=IFERROR(I17/J17,0)"
    ws["H18"] = "Cash Expense";       ws["I18"] = None; ws["I18"].fill = YELLOW
    ws["H19"] = "Net ";               ws["I19"] = "=SUM(I17:I18)"

    # Reference figure only — deliberately not deducted from Gross or Net.
    # RTO'd stock comes back sellable, so expensing it would write off
    # inventory still on hand. Rate per RTO = average product cost per piece
    # off the ProductCost TOTAL row, so it tracks the actual product mix.
    ws["H21"] = "RTO Cost";           ws["I21"] = "=(E8+E9)*M22"
    ws["M21"] = "Rate per RTO"
    ws["M22"] = (f"=IFERROR(ProductCost!L{pc_total_row}"
                 f"/ProductCost!J{pc_total_row},0)")

    # NEW: auto-filled total sales fields
    ws["H24"] = "Totale Sales";              ws["I24"] = totals["total_sales_raw"]
    ws["H25"] = "After Cancellletion sales"; ws["I25"] = totals["after_cancellation_sales"]
    ws["H26"] = "Meta spend";                ws["I26"] = "=-I7"
    ws["H27"] = "ROAS";                      ws["I27"] = "=IFERROR(I24/I26,0)"
    ws["H28"] = "After Cancellletion ROAS";  ws["I28"] = "=IFERROR(I25/I26,0)"

    for col in "DHIJKLMN":
        ws.column_dimensions[col].width = 22


# ----- main -------------------------------------------------------- #

def main():
    here = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser()
    ap.add_argument("--store", required=True,
                    help="Store identifier (matches stores/<STORE>.json)")
    ap.add_argument("--inputs-dir", default=".",
                    help="Folder containing shopify_orders.csv, bluedart*.xlsx, expressfly.csv")
    ap.add_argument("--output-dir", default=str(here / "output"))
    ap.add_argument("--stores-dir", default=str(here / "stores"))
    ap.add_argument("--month", help="Override month name (e.g. APRIL)")
    ap.add_argument("--year", type=int, help="Override year")
    args = ap.parse_args()

    cfg = load_store_config(args.store, args.stores_dir)
    store_name = cfg.get("store_name", args.store).upper()

    indir = Path(args.inputs_dir)
    outdir = Path(args.output_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    shopify_csv = indir / "shopify_orders.csv"
    expressfly_csv = indir / "expressfly.csv"
    bluedart_files = sorted(indir.glob("bluedart*.xlsx"))
    delivery_files = sorted(indir.glob("delivery*.xlsx"))

    for path, label in [(shopify_csv, "shopify_orders.csv"),
                        (expressfly_csv, "expressfly.csv")]:
        if not path.exists():
            sys.exit(f"Missing input: {path}")
    if not bluedart_files:
        sys.exit(f"No bluedart*.xlsx files in {indir}")

    print(f"[{store_name}] inputs:")
    print(f"  shopify   : {shopify_csv.name}")
    print(f"  expressfly: {expressfly_csv.name}")
    print(f"  bluedart  : {[p.name for p in bluedart_files]}")
    print(f"  delivery  : {[p.name for p in delivery_files] or 'none'}")

    # Read data
    shopify_df_raw = pd.read_csv(shopify_csv, dtype=str)
    orders_df, raw_fulfilled, totals = build_orders(shopify_csv)
    bd_rows = build_bluedart_status(bluedart_files)
    expressfly_rows = build_other_status(expressfly_csv)
    delivery_rows = build_delivery_status(delivery_files) if delivery_files else []
    other_rows = merge_other_status(expressfly_rows, delivery_rows)
    if delivery_files:
        print(f"  delivery rows: {len(delivery_rows)}"
              f" (Other status rows after merge: {len(other_rows)})")

    # Register products seen this month but missing from the store config
    seen_products = {
        parent_product(v) for v in raw_fulfilled["Lineitem name"].dropna().tolist()
    }
    new_products = register_new_products(args.store, args.stores_dir, seen_products)
    if new_products:
        print(f"  new products added to stores/{args.store}.json (cost = null):")
        for p in new_products:
            print(f"    - {p}")
        cfg = load_store_config(args.store, args.stores_dir)

    total_orders = shopify_df_raw["Name"].apply(strip_hash).dropna().nunique()
    fulfilled_orders = orders_df["Name"].dropna().nunique()

    # Month/year
    if args.month and args.year:
        month_name = args.month.upper()
        year = args.year
    else:
        m, y = detect_month_year(shopify_df_raw)
        month_name = calendar.month_name[m].upper()
        year = y

    # Classify orders into delivered/RTO buckets (with the "paid + no status" rule)
    classification = classify_orders(orders_df, bd_rows, other_rows, cfg["product_costs"])
    assumed = classification["paid_no_status_assumed_delivered"]
    print(f"  paid orders with no delivery status -> marked Delivered (Other): {len(assumed)}")

    # Order Name -> Total
    name_to_total = {}
    for _, row in orders_df.iterrows():
        n = row.get("Name")
        t = row.get("Total")
        if n is None or pd.isna(t):
            continue
        name_to_total.setdefault(n, float(t))

    def sum_amt(order_set):
        return round(sum(name_to_total.get(n, 0.0) for n in order_set), 2)

    totals_cls = {
        "total_orders": total_orders,
        "fulfilled_orders": fulfilled_orders,
        "bd_delivered": len(classification["bd_delivered"]),
        "bd_rto": len(classification["bd_rto"]),
        "other_delivered": len(classification["other_delivered"]),
        "other_rto": len(classification["other_rto"]),
        "unknown": len(classification["unknown"]),
        "bd_delivered_amt": sum_amt(classification["bd_delivered"]),
        "bd_rto_amt": sum_amt(classification["bd_rto"]),
        "other_delivered_amt": sum_amt(classification["other_delivered"]),
        "other_rto_amt": sum_amt(classification["other_rto"]),
        "unknown_amt": sum_amt(classification["unknown"]),
        "prepaid_count": int((orders_df["Financial Status"] == "paid").sum()),
        "prepaid_amt": round(
            float(pd.to_numeric(
                orders_df.loc[orders_df["Financial Status"] == "paid", "Total"],
                errors="coerce",
            ).fillna(0).sum()),
            2,
        ),
    }

    # ProductCost
    pc_data = build_product_cost(raw_fulfilled, classification, cfg["product_costs"])
    rollup_products = [p for p, *_ in pc_data["rollup"]]
    rollup_len = len(rollup_products)
    pc_total_row = 3 + rollup_len  # row containing TOTAL

    # Product-wise P&L (delivered orders only)
    pnl_rows = build_product_pnl(raw_fulfilled, classification, cfg["product_costs"])

    # Prepaid vs COD delivery/RTO behaviour
    payment_split = build_payment_split(orders_df, classification)

    # Build workbook
    wb = Workbook()
    ws_o = wb.active; ws_o.title = "Orders"
    ws_b = wb.create_sheet("BluedartSatus")
    ws_x = wb.create_sheet("OtherStatus")
    ws_p = wb.create_sheet("ProductCost")
    ws_d = wb.create_sheet("DeliveryAnalysis")
    ws_r = wb.create_sheet("Reconciliation")
    ws_pnl = wb.create_sheet("PNL")

    write_orders(ws_o, orders_df, classification)
    write_bluedart_status(ws_b, bd_rows)
    write_other_status(ws_x, other_rows, assumed)
    write_product_cost(ws_p, pc_data)
    write_delivery_analysis(ws_d, rollup_len)
    populate_unit_counts(ws_d, raw_fulfilled, rollup_products, classification)
    write_reconciliation(ws_r, totals, totals_cls, orders_df, pc_total_row,
                         payment_split)
    write_pnl(ws_pnl, pnl_rows)

    out_path = outdir / f"{month_name}_{year}_{store_name}_ORDERS.xlsx"
    wb.save(out_path)

    # Summary
    print(f"\n[{store_name}] summary:")
    print(f"  total orders               : {total_orders}")
    print(f"  fulfilled orders           : {fulfilled_orders}")
    print(f"  Bluedart delivered         : {totals_cls['bd_delivered']}")
    print(f"  Other delivered            : {totals_cls['other_delivered']}"
          f"  (incl {len(assumed)} paid-no-status assumed-delivered)")
    print(f"  Bluedart RTO               : {totals_cls['bd_rto']}")
    print(f"  Other RTO                  : {totals_cls['other_rto']}")
    print(f"  Unknown / in transit       : {totals_cls['unknown']}")
    for label in ("Prepaid", "COD"):
        d = payment_split[label]
        settled = d["delivered"] + d["rto"]
        pct = d["rto"] / settled * 100 if settled else 0
        print(f"  {label:<8} delivered {d['delivered']:>4}  RTO {d['rto']:>4}"
              f"  unknown {d['unknown']:>3}  RTO% (settled) {pct:.2f}")
    print(f"  Total Sales (raw)          : {totals['total_sales_raw']}")
    print(f"  After Cancellation Sales   : {totals['after_cancellation_sales']}")
    print(f"\nWrote: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    main()
